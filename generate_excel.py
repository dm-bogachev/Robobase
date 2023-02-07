
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()
from database.models import *
from openpyxl import Workbook

wb = Workbook()
ws = wb.active


robots = Robot.objects.all()

robot_data = []
robot_data.append("ID")
robot_data.append("Deleted?")
robot_data.append("Name")
robot_data.append("Model")
robot_data.append("Model S/N")
robot_data.append("Controller")
robot_data.append("Controller S/N")
robot_data.append("Client")
robot_data.append("Integrator")
robot_data.append("Description")
robot_data.append("Shipping date")
for j, value in enumerate(robot_data):
    cell = ws.cell(row=1, column=j+1)
    cell.value = value

for i, robot in enumerate(robots):
        data = []
        data.append(robot.pk)
        data.append(robot.deleted)
        data.append(robot.name)
        data.append(robot.arm.name)
        data.append(robot.arm_sn)
        data.append(robot.controller.name)
        data.append(robot.controller_sn)
        data.append(robot.client.name)
        if (robot.integrator is not None):
            data.append(robot.integrator.name)
        else:
            data.append(None)

        files = RobotFile.objects.all()
        append = False
        for file in files:
            if file.robot == robot:
                import os.path
                extension = os.path.splitext(file.file.path)[1][1:]
                if extension == "as":
                    append = file.file.path
        data.append(append)
        data.append(robot.description)
        data.append(robot.shipping_date)
        for j, value in enumerate(data):
            cell = ws.cell(row=i+2, column=j+1)
            cell.value = value
        pass
wb.save('test.xlsx')
